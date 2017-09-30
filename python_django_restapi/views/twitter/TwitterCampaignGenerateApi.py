import itertools, os
import traceback
from rest_framework.generics import CreateAPIView
from openpyxl import Workbook
from rest_framework.response import Response
from rest_framework import status
from datetime import datetime
import pytz
from restapi.models.twitter.TwitterRevmap import TwitterRevmap
from restapi.models.twitter.TwitterCampaign import TwitterCampaign
from restapi.models.twitter.TwitterLineItem import TwitterLineItem
from restapi.models.twitter.TwitterTargetingModels import TwitterTargeting, TwitterDevice, TwitterLocation, TwitterOsVersion
from restapi.models.twitter.TwitterPromotedTweet import TwitterPromotedTweet
from restapi.models.Campaign import Campaign
from django.utils.http import int_to_base36, base36_to_int
from restapi.models.twitter.TwitterConfig import *
import base64
from restapi.serializers.twitter.TwitterCampaignSerializer import TwitterCampaignSerializer
from restapi.serializers.CampaignSerializer import CampaignSerializer


heading = [
    'Campaign ID',
    'Funding Source ID',
    'Campaign Name',
    'Campaign Start Date',
    'Campaign End Date',
    'Campaign Status',
    'Delivery',
    'Campaign Total Budget',
    'Campaign Daily Budget',
    'Agency Credit Line Purchase Order Number',
    'Campaign Frequency Cap',
    'Campaign Frequency Cap Time Duration',
    'Ad Group ID',
    'Campaign Objective',
    'Ad Group Name',
    'Ad Group Start Time',
    'Ad Group End Time',
    'Ad Group Serving Status',
    'Ad Group Total Budget',
    'Optimization Preference',
    'Ad Group Placement',
    'Promoted Product Type',
    'Website Conversion Tag ID',
    'Charge By',
    'Advertiser Domain',
    'Bid Type',
    'Bid Amount',
    'Bid Pricing Unit',
    'Disclosure Type',
    'Disclosure Text',
    'IAB Categories',
    'Apps',
    'Amplify Programs',
    'Gender',
    'Age',
    'Location',
    'Follower Targeting',
    'Follower Targeting Lookalikes',
    'Platform',
    'User OS Version',
    'User Device',
    'Wi-Fi Only',
    'Language',
    'User Interests',
    'Mobile Carriers',
    'Device Activation Duration',
    'Exact Keywords',
    'Exact Keywords Excluded',
    'Broad Match Keywords',
    'Broad Match Keywords Excluded',
    'Unordered Keywords',
    'Unordered Keywords Excluded',
    'Phrase Keywords',
    'Phrase Keywords Excluded',
    'Tailored Audience List',
    'Tailored Audience List Excluded',
    'Tailored Audience Lookalike',
    'Behavioral Audiences',
    'Behavioral Audiences Excluded',
    'Behavioral Audiences Lookalikes',
    'Tailored Audiences from Mobile Apps',
    'Tailored Audiences from Mobile Apps Excluded',
    'Tailored Audiences from Mobile Apps Lookalikes',
    'Tailored Audiences Website Visitors',
    'Tailored Audiences Website Visitors Excluded',
    'Tailored Audiences Website Visitors Lookalikes',
    'Installed App Categories',
    'Installed App Category Lookalikes',
    'TAP - Excluded Apps',
    'TAP - Publisher App Category',
    'TV Genres',
    'TV Channels',
    'TV Shows',
    'TV Shows Airing',
    'Event Targeting',
    'Campaign Retargeting',
    'Campaign Retargeting Lookalikes',
    'Organic Tweet Retargeting',
    'Organic Tweet Retargeting Lookalikes',
    'Retargeting Engagement Type',
    'Tweet IDs',
    'Scheduled Tweet IDs',
    'Promoted Accounts ID',
    'TAP Media Creative IDs',
    'TAP Media Creative App ID',
    'TAP Media Creative Landing URL',
    'Media Creative IDs',
    'Reserved',
    '3rd party tracking with DoubleClick',
    'Reserved (2)'
]

PLATFORM_IDS = {
    'IOS': 0,
    'ANDROID': 1,
    'BLACKBERRY': 2, 
    'DESKTOP': 3, 
    'OTHER_MOBILE': 4
}

class TwitterCampaignGenerateApi(CreateAPIView):
    def _get_line_item_and_targeting(self, r, campaign_data, campaign_id_base36):
        line_item_data = campaign_data.copy()
        line_item_data['campaign_id'] = campaign_id_base36
        line_item_data['bid_amount_local_micro'] = long(round(float(self.data.get('bid_amount_default')), 2) * 1000000)
        line_item_data['name'] = r[14]
        line_item_data['objective'] = self.data['objective']
        if self.data.get('conversion_event'):
            line_item_data['primary_web_event_tag'] = self.data['conversion_event']
        if self.data.get('objective') == 'WEBSITE_CLICKS':
            line_item_data['optimization'] = 'WEBSITE_CONVERSIONS'
        line_item_data['bid_type'] = self.bid_type
        #line_items.append(line_item_data)
        targeting_data = []
        for (tw_targeting_type, cols) in TW_TARGETING_ID_EXCEL_COL.iteritems():
            if tw_targeting_type != 18:
                cols = [cols]
            for col in cols:
                if not r[col]:
                    continue
                targeting_values = r[col].split(';')
                for t in targeting_values:
                    if not t:
                        continue
                    if ':' in t:
                        targeting_value = int((t.split(':')[0])[1:])
                    else:
                        targeting_value = t
                        if tw_targeting_type == 23:
                            targeting_value = targeting_value[1:]
                    if tw_targeting_type == 9:
                        targeting_value = self.data['handle_name_id_map'][targeting_value]

                    # WI_FI only
                    if tw_targeting_type == 5:
                        targeting_value = 1

                    # platform
                    if tw_targeting_type == 3:
                        targeting_value = PLATFORM_IDS[targeting_value]

                    # location
                    if tw_targeting_type == 2:
                        loc = TwitterLocation.objects_raw.get(export_id=targeting_value)
                        targeting_value = loc.tw_targeting_id
                    
                    targeting = {
                                'name': line_item_data['name'], # this will be used as key to set tw_line_item_id
                                'targeting_value': targeting_value,
                                'tw_targeting_type': tw_targeting_type
                            }
                    # if TAILORED_AUDIENCE  
                    if tw_targeting_type == 18:
                        # Exclude?
                        if col in [55, 61]:
                            targeting['operator_type'] = 'NE'
                        else:
                            targeting['operator_type'] = 'EQ'
                    
                    targeting_data.append(targeting)

        return line_item_data, targeting_data

    def _revise_targeting_data_with_line_items_res(self, line_items_res, targeting_data):
        for line_item_res in line_items_res['data']:
            tw_line_item_id = base36_to_int(line_item_res['id'])
            line_item_name = line_item_res['name']
            for t in targeting_data:
                if t['name'] == line_item_name:
                    t['tw_line_item_id'] = tw_line_item_id
            tweets_data = {
                'line_item_id': tw_line_item_id,
                'tweet_ids': ','.join(self.data.get('tweet_ids', []))
            }
            promoted_res = TwitterPromotedTweet.set_promoted_tweet(tweets_data, self.account_id)
            if not promoted_res['success']:
                raise Exception("Associating promotable tweets with line item \"%s\" was failed..." % line_item_name)


            line_item = TwitterLineItem.objects_raw.get(pk=tw_line_item_id)
            revmap, created = TwitterRevmap.objects_raw.get_or_create(
                        **{k: v for k, v in (('campaign_id', line_item.tw_campaign_id.campaign_id),
                                             ('tw_campaign_id', line_item.tw_campaign_id),
                                             ('tw_line_item_id', line_item)) if v}
                    )
            revmap.opt_type = 'install'
            revmap.opt_value = self.cpi_target_goal
            revmap.save()

    def _generate_campaign_explode_by_campaign(self):
        campaigns_data = list()
        for r in self.rows:
            campaign_data = dict()
            campaign_data['account_id'] = self.account_id
            campaign_data['daily_budget_amount_local_micro'] = \
                long(round(float(self.data['daily_budget_amount_local_micro']), 2) * 1000000)
            campaign_data['funding_instrument_id'] = self.data['funding_instrument']
            campaign_data['name'] = r[2]
            campaign_data['paused'] = r[5] == 'PAUSED'
            campaign_data['start_time'] = self.data.get('flight_start_date', datetime.now(pytz.timezone('US/Central')).strftime('%Y-%m-%dT%H:%M:%S%z'))
            campaign_data['end_time'] = self.data.get('flight_end_date')
            campaign_data['total_budget_amount_local_micro'] = \
                long(round(float(self.data.get('total_budget_amount_local_micro') or 0), 2) * 1000000)
            campaigns_data.append(campaign_data)
        campaigns_res = TwitterCampaign.batch_create(campaigns_data, self.account_id)
        campaign_data = campaigns_data[0]
        if campaigns_res['success']:
            campaign_id = self.data.get('internal_campaign_id')
            campaign = Campaign.objects_raw.filter(pk=campaign_id).first()
            if campaign:
                self.res_json = CampaignSerializer(campaign).data
            line_items = []
            targeting_data = []
            for campaign_res in campaigns_res['data']:
                for r in self.rows:
                    if r[2] == campaign_res['name']:
                        line_item_data, targetings = self._get_line_item_and_targeting(r, campaign_data, campaign_res['id'])
                        line_items.append(line_item_data)
                        targeting_data += targetings
            # create batch line items
            line_items_res = TwitterLineItem.batch_create(line_items, self.account_id)
            if line_items_res['success']:
                self._revise_targeting_data_with_line_items_res(line_items_res, targeting_data)
                # set batch targetings
                targetings_res = TwitterTargeting.set_targeting(targeting_data, self.account_id)
                if not targetings_res['success']:
                    raise Exception(targetings_res['error']['messages'])

            else:
                raise Exception(line_items_res.get('error', {}).get('messages', ''))
        else:
            raise Exception(campaigns_res.get('error', {}).get('messages', ''))

    def _generate_campaign_explode_by_adgroup(self):
        campaign_data = {}
        campaign_data['account_id'] = self.account_id
        campaign_data['daily_budget_amount_local_micro'] = \
            long(round(float(self.data['daily_budget_amount_local_micro']), 2) * 1000000)
        campaign_data['funding_instrument_id'] = self.data['funding_instrument']
        campaign_data['name'] = self.rows[0][2]
        campaign_data['paused'] = self.rows[0][5] == 'PAUSED'
        campaign_data['start_time'] = self.data.get('flight_start_date', datetime.now(pytz.timezone('US/Central')).strftime('%Y-%m-%dT%H:%M:%S%z'))
        campaign_data['end_time'] = self.data.get('flight_end_date')
        campaign_data['total_budget_amount_local_micro'] = \
            long(round(float(self.data.get('total_budget_amount_local_micro') or 0), 2) * 1000000)

        campaign_res = TwitterCampaign.create(campaign_data)
        if campaign_res['success']:
            tw_campaign_id = base36_to_int(campaign_res['data']['id'])
            tw_campaign = TwitterCampaign.objects_raw.filter(pk=tw_campaign_id).first()
            if tw_campaign:
                self.res_json = TwitterCampaignSerializer(tw_campaign).data

            line_items = []
            targeting_data = []
            for r in self.rows:
                line_item_data, targetings = self._get_line_item_and_targeting(r, campaign_data, campaign_res['data']['id'])
                line_items.append(line_item_data)
                targeting_data += targetings

            # create batch line items
            line_items_res = TwitterLineItem.batch_create(line_items, self.account_id)
            if line_items_res['success']:
                for line_item_res in line_items_res['data']:
                    tw_line_item_id = base36_to_int(line_item_res['id'])
                    line_item_name = line_item_res['name']
                    for t in targeting_data:
                        if t['name'] == line_item_name:
                            t['tw_line_item_id'] = tw_line_item_id
                    tweets_data = {
                        'line_item_id': tw_line_item_id,
                        'tweet_ids': ','.join(self.data.get('tweet_ids', []))
                    }
                    promoted_res = TwitterPromotedTweet.set_promoted_tweet(tweets_data, self.account_id)
                    if not promoted_res['success']:
                        return False, "Associating promotable tweets with line item \"%s\" was failed..." % line_item_name


                    line_item = TwitterLineItem.objects_raw.get(pk=tw_line_item_id)
                    revmap, created = TwitterRevmap.objects_raw.get_or_create(
                        **{k: v for k, v in (('campaign_id', line_item.tw_campaign_id.campaign_id),
                                             ('tw_campaign_id', line_item.tw_campaign_id),
                                             ('tw_line_item_id', line_item)) if v}
                    )
                    revmap.opt_type = 'install'
                    revmap.opt_value = self.cpi_target_goal
                    revmap.save()                
                self._revise_targeting_data_with_line_items_res(line_items_res, targeting_data)
                # set batch targetings
                targetings_res = TwitterTargeting.set_targeting(targeting_data, self.account_id)
                if not targetings_res['success']:
                    raise Exception(targetings_res['error']['messages'])

            else:
                raise Exception(line_items_res.get('error', {}).get('messages', ''))
        else:
            raise Exception(campaign_res['message'])

    def post(self, request, *args, **kwargs):
        self.data = request.data
        data = self.data
        self.account_id = long(data.get('account_id'))
        self.bid_type = self.data.get('pricing', 'MAX_BID').replace('_BID', '')
        self.cpi_target_goal = round(float(data.get('bid_amount_default')), 2)
        if 'TARGET' in self.bid_type:
            self.bid_type = 'TARGET'

        explode_adgroup = data.get('explode_adgroup')
        campaign_id = data['internal_campaign_id']
        campaign_name = '%s {%s}' % (data['campaign'], data['internal_campaign_id'])
        file_name = '%s_%s.xlsx' % (data['campaign'], datetime.now().strftime('%d%m%Y'))
        dir_path = os.path.join('/tmp/', 'tw_campaign_excels')
        if not os.path.exists(dir_path):
            try:
                os.makedirs(dir_path)
            except OSError as exc:
                return Response(dict(error="Can't create excel file on server!"), status=status.HTTP_400_BAD_REQUEST)

        full_path = os.path.join(dir_path, file_name)

        wb = Workbook()
        ws = wb.active
        ws.title = "Campaigns"
        for column in range(len(heading)):
            ws.cell(row=1, column=column+1).value = heading[column]

        
        if data['objective'] == 'APP_INSTALLS':
            os_name = 'IOS' if data['os'] == 1 else 'ANDROID'

            if data.get('min_version') or data.get('devices'):
                data['os'] = ''
            else:
                data['os'] = os_name

            if data.get('min_version'):
                data['min_version'] = 'i%s:%s;' % (data['min_version']['tw_targeting_id'], data['min_version']['os_version'])
            data['devices'] = ';'.join(['i%s:%s' % (d['tw_targeting_id'], d['device']) for d in data.get('devices', [])])
        
        if data['objective'] == 'WEBSITE_CLICKS':
            PLATFORMS = ['IOS', 'ANDROID', 'BLACKBERRY', 'DESKTOP', 'OTHER_MOBILE']
            _os = []
            
            for i in data['web_clicks_os']:
                _os.append(PLATFORMS[i])
            if data.get('web_click_android_min_version') or data.get('web_click_android_devices'):
                _os.pop(_os.index('ANDROID'))

            if data.get('web_click_ios_min_version') or data.get('web_click_ios_devices'):
                _os.pop(_os.index('IOS'))
            
            data['os'] = ';'.join(_os)
            os_name = '-'.join(_os)
            if data.get('web_click_ios_min_version'):
                row = TwitterOsVersion.objects_raw.get(pk=data['web_click_ios_min_version'])
                data['web_click_ios_min_version'] = 'i%s:%s' % (row.tw_targeting_id, row.os_version)
            
            if data.get('web_click_android_min_version'):
                row = TwitterOsVersion.objects_raw.get(pk=data['web_click_android_min_version'])
                data['web_click_android_min_version'] = 'i%s:%s' % (row.tw_targeting_id, row.os_version)
            
            data['min_version'] = []
            if data.get('web_click_ios_min_version'):
                data['min_version'].append(data['web_click_ios_min_version'])
            if data.get('web_click_android_min_version'):
                data['min_version'].append(data['web_click_android_min_version'])
            data['min_version'] = ';'.join(data['min_version'])

            data['devices'] = data['web_click_ios_devices'] + data['web_click_android_devices']
            if data['devices']:
                rows = TwitterDevice.objects_raw.filter(pk__in=data['devices'].split(',')).all()
                data['devices'] = ';'.join(['i%s:%s' % (r.tw_targeting_id, r.device) for r in rows])    
            else:
                data['devices'] = ''

            """
            os_version = TwitterOsVersion.objects_raw.get(pk=data['min_version']['tw_targeting_id'])
            versions = TwitterOsVersion.objects_raw.filter(platform=os_version.platform, number__gte=os_version.number).all()
            res = []
            for v in versions:
                res.append('i%s:%s' % (v.tw_targeting_id, v.number))
            data['min_version'] = ';'.join(res)
            """

        data['handles_group'] = [x for x in data['handles_group'] if x]
        data['handles_group'] = \
            [[{'id': h['text'], 'text': '@'+h['text']} for h in handles] for handles in data.get('handles_group', [])]

        #data['keywords_group'] = [x for x in data['keywords_group'] if x]
        data['keywords_group'] = \
            [[{'id': h['text'], 'text': h['text']} for h in keywords] for keywords in data.get('keywords_group', [])]

        data['tailored_audiences_list_included'] = ';'.join(['i%s:Custom audience targeting' % a['tw_targeting_id'] for a in data.get('tailored_audiences_list_included', [])])
        data['tailored_audiences_from_mobile_apps_included'] = ';'.join(['i%s:Mobile audience targeting' % a['tw_targeting_id'] for a in data.get('tailored_audiences_from_mobile_apps_included', [])])

        data['tailored_audiences_list_excluded'] = ';'.join(['i%s:Custom audience targeting' % a['tw_targeting_id'] for a in data.get('tailored_audiences_list_excluded', [])])
        data['tailored_audiences_from_mobile_apps_excluded'] = ';'.join(['i%s:Mobile audience targeting' % a['tw_targeting_id'] for a in data.get('tailored_audiences_from_mobile_apps_excluded', [])])

        data['carriers'] = \
            ';'.join(['i%s:%s' % (d['id'], d['label']) for d in data.get('carriers', [])])

        if data.get('target_each_country'):
            data['locations_group'] = [[l] for l in data.get('locations_group', [[]])[0]]

        if data.get('enable_event_targeting'):
            data['event_targeting'] = ['i%s' % t for t in data['event_targeting']]
        else:
            data['event_targeting'] = ['']

        # make ad group names
        targeting_fields = [
            {'data_field': 'locations_group', 'list_id': 'export_id', 'list_text': 'location_name'},
            {'data_field': 'handles_group', 'list_id': 'id', 'list_text': 'text'},
            {'data_field': 'keywords_group', 'list_id': 'id', 'list_text': 'text'},
            {'data_field': 'app_categories_group', 'list_id': 'id', 'list_text': 'label'},
            {'data_field': 'interests_group', 'list_id': 'id', 'list_text': 'label'}
        ]


        targeting_excel_values = {}
        for tf in targeting_fields:
            values_group = data.get(tf['data_field'], [])
            targeting_values_group = []
            for values in values_group:
                targeting_values = []
                for v in values:
                    id = v.get(tf['list_id'])
                    text = v.get(tf['list_text'])
                    # 'keyword' and 'handle' should be excluded
                    if '_id' in tf['list_id'] or tf['list_text'] == 'label':
                        targeting_values.append('i%s:%s' % (id, text))
                    else:
                        targeting_values.append(id)

                if targeting_values:
                    targeting_values_group.append(';'.join(targeting_values))

            targeting_excel_values[tf['data_field']] = targeting_values_group if targeting_values_group else ['']

        conversion_event = data.get('conversion_event', '')
        if conversion_event:
            conversion_event = "i%d" % base36_to_int(conversion_event)
        rows = list()
        #'Campaign ID' 0
        rows.append([''])
        #'Funding Source ID' 1
        rows.append(['i%s' % int(data['funding_instrument'], 36)])
        #'Campaign Name', 2
        rows.append([campaign_name])
        #'Campaign Start Date' 3
        rows.append([data.get('flight_start_date', datetime.now().strftime('%d-%b-%Y %H:%M'))])
        #'Campaign End Date' 4
        rows.append([data.get('flight_end_date', '')])
        #'Campaign Status' 5
        rows.append([data['status']])
        #'Delivery', 6
        rows.append(['TRUE'])
        #'Campaign Total Budget' 7
        rows.append([round(float(data.get('total_budget_amount_local_micro') or 0), 2)])
        #'Campaign Daily Budget' 8
        rows.append([round(float(data['daily_budget_amount_local_micro']), 2)])
        #'Agency Credit Line Purchase Order Number', 9
        rows.append([''])
        #'Campaign Frequency Cap' 10
        rows.append([''])
        #'Campaign Frequency Cap Time Duration' 11
        rows.append([''])
        #'Ad Group ID', 12
        rows.append([''])
        #'Campaign Objective', 13
        rows.append([data.get('objective')])
        #'Ad Group Name', 14
        rows.append([''])
        #'Ad Group Start Time', 15
        rows.append([''])
        #'Ad Group End Time', 16
        rows.append([''])
        #'Ad Group Serving Status', 17
        rows.append([data['status']])
        #'Ad Group Total Budget', 18
        rows.append([''])
        #'Optimization Preference', 19
        rows.append(['NONE' if data['objective'] == 'APP_INSTALLS' else 'CONVERSIONS'])
        # 'Ad Group Placement', 20
        temp = []
        if data.get('tweets_on_profiles'):
            temp.append('PROFILES')
        if data.get('tweets_on_timeline'):
            temp.append('TIMELINES')
        rows.append([';'.join(temp)])
        # 'Promoted Product Type', 21
        rows.append(['PROMOTED_TWEETS'])
        
        # 'Website Conversion Tag ID', 22
        rows.append([conversion_event])
        
        # 'Charge By', 23
        rows.append(['-'])
        # 'Advertiser Domain', 24
        rows.append(['-'])
        # 'Bid Type', 25
        rows.append([data.get('pricing')])
        # 'Bid Amount', 26
        rows.append([float(data.get('bid_amount_default'))])
        # 'Bid Pricing Unit', 27
        rows.append([''])
        # 'Disclosure Type', 28
        rows.append(['NONE'])
        # 'Disclosure Text', 29
        rows.append([''])
        # 'IAB Categories' 30
        rows.append([''])
        # 'Apps', 31
        rows.append([''])
        # 'Amplify Programs', 32
        rows.append([''])
        # 'Gender', 33
        rows.append([''])
        # 'Age', 34
        rows.append([''])
        # 'Location', 35
        rows.append(targeting_excel_values.get('locations_group', ['']))
        # 'Follower Targeting', 36
        rows.append([''])
        # 'Follower Targeting Lookalikes', 37
        rows.append(targeting_excel_values.get('handles_group', ['']))
        # 'Platform', 38
        rows.append([data['os']])
        # 'User OS Version', 39
        rows.append([data.get('min_version', '')])
        # 'User Device', 40
        rows.append([data.get('devices', '')])
        # 'Wi-Fi Only', 41
        rows.append(['True' if data.get('wifi_only') else ''])
        # 'Language', 42
        rows.append(['en'])
        # 'User Interests', 43
        rows.append(targeting_excel_values.get('interests_group', ['']))
        # 'Mobile Carriers', 44
        rows.append([data.get('carriers', '')])
        # 'Device Activation Duration', 45
        rows.append([''])
        # 'Exact Keywords', 46
        rows.append([''])
        # 'Exact Keywords Excluded', 47
        rows.append([''])
        # 'Broad Match Keywords', 48
        rows.append(targeting_excel_values.get('keywords_group', ['']))
        # 'Broad Match Keywords Excluded', 49
        rows.append([''])
        # 'Unordered Keywords', 50
        rows.append([''])
        # 'Unordered Keywords Excluded', 51
        rows.append([''])
        # 'Phrase Keywords', 52
        rows.append([''])
        # 'Phrase Keywords Excluded', 53
        rows.append([''])
        # 'Tailored Audience List', 54
        rows.append([data.get('tailored_audiences_list_included', '')])
        # 'Tailored Audience List Excluded', 55
        rows.append([data.get('tailored_audiences_list_excluded', '')])
        # 'Tailored Audience Lookalike', 56
        rows.append([''])
        # 'Behavioral Audiences', 57
        rows.append([''])
        # 'Behavioral Audiences Excluded', 58
        rows.append([''])
        # 'Behavioral Audiences Lookalikes', 59
        rows.append([''])
        # 'Tailored Audiences from Mobile Apps', 60
        rows.append([data.get('tailored_audiences_from_mobile_apps_included', '')])
        # 'Tailored Audiences from Mobile Apps Excluded', 61
        rows.append([data.get('tailored_audiences_from_mobile_apps_excluded', '')])
        # 'Tailored Audiences from Mobile Apps Lookalikes', 62
        rows.append([''])
        # 'Tailored Audiences Website Visitors', 63
        rows.append([''])
        # 'Tailored Audiences Website Visitors Excluded', 64
        rows.append([''])
        # 'Tailored Audiences Website Visitors Lookalikes', 65
        rows.append([''])
        # 'Installed App Categories', 66
        rows.append(targeting_excel_values.get('app_categories_group', ['']))
        # 'Installed App Category Lookalikes', 67
        rows.append([''])
        # 'TAP - Excluded Apps', 68
        rows.append([''])
        # 'TAP - Publisher App Category', 69
        rows.append([''])
        # 'TV Genres', 70
        rows.append([''])
        # 'TV Channels', 71
        rows.append([''])
        # 'TV Shows', 72
        rows.append([''])
        # 'TV Shows Airing', 73
        rows.append([''])
        # 'Event Targeting', 74
        rows.append(data.get('event_targeting', ['']))
        # 'Campaign Retargeting', 75
        rows.append([''])
        # 'Campaign Retargeting Lookalikes', 76
        rows.append([''])
        # 'Organic Tweet Retargeting', 77
        rows.append([''])
        # 'Organic Tweet Retargeting Lookalikes', 78
        rows.append([''])
        # 'Retargeting Engagement Type', 79
        rows.append([''])
        # 'Tweet IDs', 80
        rows.append([';'.join(['i%s' % t for t in data.get('tweet_ids', [])])])
        # 'Scheduled Tweet IDs', 81
        rows.append([''])
        # 'Promoted Accounts ID', 82
        rows.append([''])
        # 'TAP Media Creative IDs', 83
        rows.append([''])
        # 'TAP Media Creative App ID', 84
        rows.append([''])
        # 'TAP Media Creative Landing URL', 85
        rows.append([''])
        # 'Media Creative IDs', 86
        rows.append([''])
        # 'Reserved', 87
        rows.append([''])
        # '3rd party tracking with DoubleClick', 88
        rows.append([''])
        # 'Reserved (2)' 89
        rows.append([''])
        row_index = 2

        _handles_group = targeting_excel_values.get('handles_group')
        _keywords_group = targeting_excel_values.get('keywords_group')

        if not data['separate_keyword_follower'] or not _handles_group[0] or not _keywords_group[0]:
            rows = list(itertools.product(*rows))
        else:
            temp = []
            rows[48] = ['']
            temp = list(itertools.product(*rows))
            rows[37] = ['']
            rows[48] = _keywords_group
            temp = temp + list(itertools.product(*rows))
            rows = temp

        tv_info = data.get('tv_info', {})
        behavior_info = data.get('behavior_info', {})

        if len(rows) > 100:
            return Response(data=dict(message='Please be aware that you may create up to 100 total Ad Groups per campaign!'),
                        status=status.HTTP_400_BAD_REQUEST)

        final_rows = []
        for row in rows:
            row_list = list(row)
            location_name = row_list[35] + '_' if row_list[35] else ''
            suffix = '%s_%s' % (location_name.split(':', 1)[-1], os_name)
            row_list[14] = 'Ad Group %s_%s' % (suffix, row_index-1)

            #TV Marketing
            _countries = row_list[35].split(';')

            if row_list[35] and len(_countries) == 1:
                _country = _countries[0].split(':')[1]
                _tv_info = tv_info.get(_country, False)
                if not _tv_info and _country == 'United States':
                    _tv_info = tv_info.get('United States - Hispanic', False)
                if _tv_info:
                    #TV Genres
                    row_list[70] = ';'.join(_tv_info.get('genres', []))

                    #TV Channels
                    row_list[71] = ';'.join(_tv_info.get('channels', '').split(','))

                    #TV Shows
                    row_list[72] = ';'.join(_tv_info.get('shows', '').split(','))

                _behavior_info = behavior_info.get(_country, False)
                if _behavior_info:
                    row_list[59] = ';'.join(['i%s:Partner audience lookalike targeting' % info for info in _behavior_info])

            if explode_adgroup:
                row_list[0] = 'new C1'
            else:
                row_list[0] = 'new C%s' % (row_index - 1)
                row_list[2] = '%s_%s_%s {%s}' % (data['campaign'], suffix, row_index - 1, data['internal_campaign_id'])

            col_index = 1
            for c in row_list:
                ws.cell(row=row_index, column=col_index).value = c
                col_index += 1

            final_rows.append(row_list)
            row_index += 1

        wb.save(full_path)

        self.rows = final_rows
        self.res_json = {}
        api_res_status = {}
        message  = ''
        success = True
        if data.get('create_tw_campaign'):
            try:
                if data.get('explode_adgroup'):
                    self._generate_campaign_explode_by_adgroup()
                else:
                    self._generate_campaign_explode_by_campaign()
            except Exception as e:
                traceback.print_exc()
                success = False
                message = str(e)

            api_res_status['success'] = success
            api_res_status['campaign'] = self.res_json
            api_res_status['message'] = message
            api_res_status['line_item_count'] = len(rows)

        excel_file = open(full_path, "rb")
        encoded_string = base64.b64encode(excel_file.read())
        return Response(dict(results=dict(api_status=api_res_status, base64_file_content=encoded_string, campaign_id=campaign_id,
                                          file_name=file_name)))